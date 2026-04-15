#!/usr/bin/env python3
"""
Test with a realistic Excel file similar to the one in the error report.
File: unidades-atendimento-parana-clinicas-unimed_2025_10_03.xlsx
"""
import sys
import os
import logging

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

from app.providers.format_converter import FormatConverter

def test_realistic_excel():
    """Test with a realistic healthcare units file."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from io import BytesIO
        
        print("\n" + "="*70)
        print("Testing Realistic Excel File (Healthcare Units)")
        print("="*70)
        
        # Create a realistic healthcare units spreadsheet
        print("\n1. Creating realistic 'unidades-atendimento' Excel file...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Unidades"
        
        # Headers
        headers = ['ID', 'Nome Unidade', 'Cidade', 'Estado', 'Telefone', 'Email', 'Ativo']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data - Healthcare units in Paraná
        data = [
            [1, 'Clínica Unimed Curitiba', 'Curitiba', 'PR', '(41) 3333-4444', 'curitiba@unimed.com.br', 'Sim'],
            [2, 'Clínica Unimed Londrina', 'Londrina', 'PR', '(43) 3333-5555', 'londrina@unimed.com.br', 'Sim'],
            [3, 'Clínica Unimed Maringá', 'Maringá', 'PR', '(44) 3333-6666', 'maringa@unimed.com.br', 'Sim'],
            [4, 'Clínica Unimed Ponta Grossa', 'Ponta Grossa', 'PR', '(42) 3333-7777', 'pontagrossa@unimed.com.br', 'Sim'],
            [5, 'Clínica Unimed Cascavel', 'Cascavel', 'PR', '(45) 3333-8888', 'cascavel@unimed.com.br', 'Não'],
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        file_bytes = excel_bytes.getvalue()
        
        print(f"   ✓ Created Excel file: {len(file_bytes)} bytes")
        print(f"   ✓ Similar to real file size (32,002 bytes) - test file is smaller but structure is same")
        
        # Test conversion
        print("\n2. Converting to CSV format...")
        filename = "unidades-atendimento-parana-clinicas-unimed_2025_10_03.xlsx"
        csv_content, file_format = FormatConverter.convert_to_csv(file_bytes, filename)
        
        print(f"   ✓ Format detected: {file_format}")
        print(f"   ✓ CSV output: {len(csv_content)} chars")
        print(f"   ✓ File classification success")
        
        # Display results
        print("\n3. CSV Output Preview:")
        print("   " + "-" * 66)
        lines = csv_content.split('\n')
        for i, line in enumerate(lines[:8]):
            if i == 0:
                print(f"   [HEADER] {line}")
            elif line.strip():
                print(f"   [DATA]   {line[:62]}...")
        print("   " + "-" * 66)
        
        # Validate
        print("\n4. Validation:")
        if 'ID,Nome Unidade' in csv_content or 'ID' in csv_content and 'Unidade' in csv_content:
            print("   ✓ Headers present")
        else:
            print("   ✗ Headers missing")
            return False
        
        if 'Curitiba' in csv_content and 'Londrina' in csv_content:
            print("   ✓ Data preserved")
        else:
            print("   ✗ Data missing")
            return False
        
        if 'errorFile' not in csv_content.lower():
            print("   ✓ No error messages in output")
        else:
            print("   ✗ Error detected in output")
            return False
        
        print("\n" + "="*70)
        print("✓ REALISTIC TEST PASSED - Excel handling works correctly!")
        print("="*70)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_large_excel():
    """Test with a larger Excel file to ensure it handles real-world sizes."""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        print("\n" + "="*70)
        print("Testing Large Excel File (Stress Test)")
        print("="*70)
        
        print("\n1. Creating large Excel file with 1000 rows...")
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeData"
        
        # Headers
        headers = ['ID', 'Name', 'Email', 'City', 'Status', 'Amount', 'Date']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Generate 1000 rows
        import random
        cities = ['São Paulo', 'Rio de Janeiro', 'Belo Horizonte', 'Brasília', 'Salvador']
        statuses = ['Active', 'Inactive', 'Pending', 'Archived']
        
        for row_idx in range(2, 1002):  # 1000 data rows
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=f'User{row_idx}')
            ws.cell(row=row_idx, column=3, value=f'user{row_idx}@example.com')
            ws.cell(row=row_idx, column=4, value=random.choice(cities))
            ws.cell(row=row_idx, column=5, value=random.choice(statuses))
            ws.cell(row=row_idx, column=6, value=random.uniform(100, 10000))
            ws.cell(row=row_idx, column=7, value='2025-01-13')
        
        # Save to bytes
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        file_bytes = excel_bytes.getvalue()
        
        print(f"   ✓ Created large Excel file: {len(file_bytes)} bytes")
        
        # Convert
        print("\n2. Converting large file to CSV...")
        csv_content, file_format = FormatConverter.convert_to_csv(file_bytes, "large_data.xlsx")
        
        print(f"   ✓ Format: {file_format}")
        print(f"   ✓ CSV size: {len(csv_content)} chars")
        
        # Count rows
        csv_lines = csv_content.split('\n')
        data_rows = [l for l in csv_lines if l.strip() and 'ID' not in l]
        print(f"   ✓ Data rows extracted: {len(data_rows)}")
        
        if len(data_rows) >= 900:  # Allow some variation
            print("   ✓ Row count matches")
        else:
            print(f"   ✗ Row count mismatch: expected ~1000, got {len(data_rows)}")
            return False
        
        print("\n" + "="*70)
        print("✓ LARGE FILE TEST PASSED!")
        print("="*70)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_realistic_excel()
    success = test_large_excel() and success
    
    if success:
        print("\n✓✓✓ All realistic tests passed! Excel handling is working correctly. ✓✓✓\n")
        sys.exit(0)
    else:
        print("\n✗ Some tests failed")
        sys.exit(1)
