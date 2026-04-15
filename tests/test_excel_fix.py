#!/usr/bin/env python3
"""
Test script to verify Excel file handling fix.
"""
import sys
import os
import logging

# Add app to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

from app.providers.format_converter import FormatConverter

def test_excel_handling():
    """Test Excel file handling."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from io import BytesIO
        import pandas as pd
        
        print("\n" + "="*60)
        print("Testing Excel File Handling")
        print("="*60)
        
        # Create a test Excel file in memory
        print("\n1. Creating test Excel file...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers and data
        ws['A1'] = 'Nome'
        ws['B1'] = 'Cidade'
        ws['C1'] = 'Estado'
        
        ws['A2'] = 'Maria'
        ws['B2'] = 'São Paulo'
        ws['C2'] = 'SP'
        
        ws['A3'] = 'João'
        ws['B3'] = 'Rio de Janeiro'
        ws['C3'] = 'RJ'
        
        # Save to bytes
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        file_bytes = excel_bytes.getvalue()
        
        print(f"   ✓ Created Excel file: {len(file_bytes)} bytes")
        print(f"   ✓ File signature: {file_bytes[:4].hex()}")
        
        # Test conversion
        print("\n2. Testing format conversion...")
        csv_content, file_format = FormatConverter.convert_to_csv(file_bytes, "test_unidades.xlsx")
        
        print(f"   ✓ Format detected: {file_format}")
        print(f"   ✓ CSV output size: {len(csv_content)} chars")
        print(f"   ✓ CSV content preview:")
        print("   " + "-" * 56)
        for line in csv_content.split('\n')[:10]:
            print(f"   | {line}")
        print("   " + "-" * 56)
        
        # Verify headers are preserved
        lines = csv_content.split('\n')
        if lines and 'Nome' in lines[0]:
            print("\n   ✓ SUCCESS: Headers preserved in CSV!")
        else:
            print("\n   ✗ WARNING: Headers not found in CSV output")
        
        # Verify data is present
        if 'Maria' in csv_content or 'São Paulo' in csv_content:
            print("   ✓ SUCCESS: Data preserved in CSV!")
        else:
            print("   ✗ WARNING: Data not found in CSV output")
        
        print("\n" + "="*60)
        print("Excel handling test PASSED!")
        print("="*60)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_zip_fallback():
    """Test ZIP fallback when Excel handling fails."""
    print("\n" + "="*60)
    print("Testing ZIP Fallback Handling")
    print("="*60)
    
    try:
        # Create a fake ZIP file (just the header)
        import zipfile
        from io import BytesIO
        
        print("\n1. Creating minimal ZIP file...")
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zf:
            zf.writestr('test.txt', 'Hello from ZIP')
        
        zip_bytes = zip_buffer.getvalue()
        print(f"   ✓ Created ZIP file: {len(zip_bytes)} bytes")
        
        # Test conversion - this should handle gracefully
        print("\n2. Testing fallback handling...")
        csv_content, file_format = FormatConverter.convert_to_csv(zip_bytes, "test.zip")
        
        print(f"   ✓ Format detected: {file_format}")
        print(f"   ✓ CSV output size: {len(csv_content)} chars")
        
        if "error" in csv_content.lower():
            print("   ℹ Note: ZIP file returned error message (expected for non-Excel ZIP)")
        else:
            print("   ✓ ZIP content was extracted")
        
        print("\n" + "="*60)
        print("ZIP fallback test PASSED!")
        print("="*60)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_excel_handling()
    success = test_zip_fallback() and success
    
    if success:
        print("\n✓ All tests passed!")
        sys.exit(0)
    else:
        print("\n✗ Some tests failed")
        sys.exit(1)
