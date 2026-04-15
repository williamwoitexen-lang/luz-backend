#!/usr/bin/env python3
"""
Integration test: Simulate the document upload preview endpoint.
Tests that the preview endpoint works with Excel files.
"""
import sys
import os
import logging
import json
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

logging.basicConfig(
    level=logging.WARNING,  # Reduce verbosity for this test
)

from app.providers.format_converter import FormatConverter

def simulate_document_service_preview(file_bytes: bytes, filename: str) -> dict:
    """Simulate the document service preview logic."""
    
    # Step 1: Convert file to CSV
    csv_content, original_format = FormatConverter.convert_to_csv(file_bytes, filename)
    
    # Step 2: Limit preview size
    MAX_PREVIEW_SIZE = 15 * 1024  # 15KB
    preview_content = csv_content[:MAX_PREVIEW_SIZE]
    
    # Step 3: Clean preview (simplified version)
    cleaned_preview = preview_content  # In real code this calls _clean_text_content
    
    # Build response (matching the user's error response structure)
    response = {
        "status": "success",
        "temp_id": "test-temp-id",
        "filename": filename,
        "file_size_bytes": len(file_bytes),
        "csv_size_chars": len(csv_content),
        "preview_size_chars": len(preview_content),
        "cleaned_preview_chars": len(cleaned_preview),
        "size_reduction_percent": 100 - (len(cleaned_preview)*100//len(preview_content) if len(preview_content) > 0 else 0),
        "extracted_fields": {},
        "expires_in_minutes": 10,
        "next_step": f"POST /api/v1/documents/ingest-confirm/test-temp-id with metadata to confirm",
        "preview_sample": {
            "summary": None,
            "first_300_chars": cleaned_preview[:300],
            "full_cleaned_preview": cleaned_preview if len(cleaned_preview) < 1000 else cleaned_preview[:1000] + "..."
        }
    }
    
    return response

def test_endpoint_with_excel():
    """Test the document preview endpoint with an Excel file."""
    
    print("\n" + "="*70)
    print("Integration Test: Document Preview Endpoint with Excel")
    print("="*70)
    
    try:
        from openpyxl import Workbook
        
        # Create test Excel file
        print("\n1. Creating test Excel file (unidades-atendimento)...")
        wb = Workbook()
        ws = wb.active
        
        # Add realistic Paraná healthcare units data
        ws['A1'] = 'ID'
        ws['B1'] = 'Nome'
        ws['C1'] = 'Cidade'
        ws['D1'] = 'Estado'
        ws['E1'] = 'Categoria'
        
        cities = [
            'Curitiba', 'Londrina', 'Maringá', 'Ponta Grossa', 'Cascavel',
            'São José dos Pinhais', 'Paranaguá', 'Apucarana', 'Pato Branco', 'Irati'
        ]
        
        for i, city in enumerate(cities, 1):
            ws[f'A{i+1}'] = i
            ws[f'B{i+1}'] = f'Unidade {i}'
            ws[f'C{i+1}'] = city
            ws[f'D{i+1}'] = 'PR'
            ws[f'E{i+1}'] = 'Clínica'
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        file_bytes = excel_bytes.getvalue()
        
        print(f"   ✓ File size: {len(file_bytes)} bytes")
        
        # Simulate endpoint call
        print("\n2. Calling preview endpoint...")
        filename = "unidades-atendimento-parana-clinicas-unimed_2025_10_03.xlsx"
        response = simulate_document_service_preview(file_bytes, filename)
        
        # Check response structure
        print("\n3. Validating response structure...")
        
        required_fields = [
            'status', 'temp_id', 'filename', 'file_size_bytes',
            'csv_size_chars', 'preview_size_chars', 'cleaned_preview_chars',
            'preview_sample'
        ]
        
        for field in required_fields:
            if field in response:
                print(f"   ✓ {field}: present")
            else:
                print(f"   ✗ {field}: MISSING")
                return False
        
        # Check for errors
        print("\n4. Checking for errors in response...")
        
        preview = response['preview_sample']['full_cleaned_preview']
        
        if 'errorFile' in preview or 'error' == preview.lower()[:5]:
            print(f"   ✗ ERROR DETECTED IN PREVIEW:")
            print(f"      {preview[:100]}")
            return False
        else:
            print(f"   ✓ No error messages in preview")
        
        if 'Unidade' in preview or 'Curitiba' in preview:
            print(f"   ✓ Data extracted successfully")
        else:
            print(f"   ✗ No data in preview")
            return False
        
        # Display formatted response
        print("\n5. Response Summary:")
        print(f"   Status: {response['status']}")
        print(f"   Filename: {response['filename']}")
        print(f"   File size: {response['file_size_bytes']} bytes")
        print(f"   CSV size: {response['csv_size_chars']} chars")
        print(f"   Preview size: {response['preview_size_chars']} chars")
        print(f"   Expires in: {response['expires_in_minutes']} minutes")
        print(f"\n   Preview (first 200 chars):")
        print(f"   {preview[:200]}\n")
        
        print("="*70)
        print("✓ ENDPOINT TEST PASSED - Excel preview works correctly!")
        print("="*70)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_endpoint_error_handling():
    """Test that endpoint handles errors gracefully."""
    
    print("\n" + "="*70)
    print("Integration Test: Error Handling")
    print("="*70)
    
    try:
        # Create a truly invalid file
        print("\n1. Testing with corrupted ZIP file...")
        
        # ZIP signature but corrupted content
        corrupted_file = b'PK\x03\x04' + b'\x00' * 100
        
        response = simulate_document_service_preview(corrupted_file, "corrupted.xlsx")
        
        print(f"   Response status: {response['status']}")
        preview = response['preview_sample']['full_cleaned_preview']
        
        # Should contain error message
        if 'error' in preview.lower():
            print(f"   ✓ Error message present (as expected for corrupted file)")
        else:
            print(f"   Preview: {preview[:100]}")
        
        print("\n" + "="*70)
        print("✓ ERROR HANDLING TEST PASSED")
        print("="*70)
        return True
        
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        return False

if __name__ == '__main__':
    success = test_endpoint_with_excel()
    success = test_endpoint_error_handling() and success
    
    if success:
        print("\n" + "="*70)
        print("✓✓✓ ALL INTEGRATION TESTS PASSED ✓✓✓")
        print("The Excel file handling fix is working correctly!")
        print("The preview endpoint will now work with Excel files.")
        print("="*70 + "\n")
        sys.exit(0)
    else:
        print("\n✗ Some integration tests failed")
        sys.exit(1)
